"""گزارش‌گیری سفارش‌ها — فقط خواندن، هرگز حذف نمی‌شود (نیاز به توکن)"""
from datetime import datetime, time, timedelta
from io import BytesIO

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from .. import models
from ..auth import require_role
from ..database import get_db
from ..schemas import OrderOut
from .orders import order_to_out

router = APIRouter(
    prefix="/api/reports",
    tags=["reports"],
    dependencies=[Depends(require_role("admin"))],
)


def _date_range(period: str, date_from: str | None, date_to: str | None):
    """محاسبه بازه زمانی بر اساس period یا تاریخ دستی"""
    now = datetime.now()
    if period == "today":
        return datetime.combine(now.date(), time.min), datetime.combine(now.date(), time.max)
    elif period == "week":
        start = now - timedelta(days=now.weekday())
        return datetime.combine(start.date(), time.min), datetime.combine(now.date(), time.max)
    elif period == "month":
        return datetime(now.year, now.month, 1), datetime.combine(now.date(), time.max)
    elif period == "year":
        return datetime(now.year, 1, 1), datetime.combine(now.date(), time.max)
    elif period == "custom" and date_from and date_to:
        try:
            start = datetime.strptime(date_from, "%Y-%m-%d")
            end = datetime.combine(datetime.strptime(date_to, "%Y-%m-%d").date(), time.max)
            return start, end
        except ValueError:
            pass
    # همه زمان‌ها
    return None, None


@router.get("/orders", response_model=list[OrderOut])
def list_all_orders(
    period: str = Query(default="all", description="today|week|month|year|custom|all"),
    date_from: str | None = Query(default=None, description="YYYY-MM-DD"),
    date_to: str | None = Query(default=None, description="YYYY-MM-DD"),
    search: str | None = Query(default=None, description="جستجو در کد، نام مشتری"),
    status: str | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=200),
    db: Session = Depends(get_db),
):
    """لیست کامل تمام سفارش‌ها با فیلتر و صفحه‌بندی — هرگز چیزی حذف نمی‌شود"""
    query = db.query(models.Order)

    start, end = _date_range(period, date_from, date_to)
    if start:
        query = query.filter(models.Order.created_at >= start)
    if end:
        query = query.filter(models.Order.created_at <= end)

    if status:
        query = query.filter(models.Order.status == status)

    if search:
        s = f"%{search}%"
        query = query.filter(
            models.Order.code.ilike(s)
            | models.Order.customer_name.ilike(s)
        )

    orders = (
        query.order_by(models.Order.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    return [order_to_out(o, with_qr=False) for o in orders]


@router.get("/orders/count")
def count_orders(
    period: str = Query(default="all"),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    status: str | None = Query(default=None),
    search: str | None = Query(default=None),
    db: Session = Depends(get_db),
):
    """تعداد کل سفارش‌ها برای صفحه‌بندی"""
    query = db.query(models.Order)
    start, end = _date_range(period, date_from, date_to)
    if start:
        query = query.filter(models.Order.created_at >= start)
    if end:
        query = query.filter(models.Order.created_at <= end)
    if status:
        query = query.filter(models.Order.status == status)
    if search:
        s = f"%{search}%"
        query = query.filter(
            models.Order.code.ilike(s) | models.Order.customer_name.ilike(s)
        )
    return {"total": query.count()}


@router.get("/summary")
def summary(
    period: str = Query(default="month"),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    db: Session = Depends(get_db),
):
    """آمار کلی: تعداد و مجموع مبلغ به تفکیک وضعیت"""
    query = db.query(models.Order)
    start, end = _date_range(period, date_from, date_to)
    if start:
        query = query.filter(models.Order.created_at >= start)
    if end:
        query = query.filter(models.Order.created_at <= end)

    orders = query.all()
    result: dict = {"total_count": len(orders), "total_amount": 0, "by_status": {}}
    for o in orders:
        result["total_amount"] += o.total_amount
        result["by_status"].setdefault(o.status, {"count": 0, "amount": 0})
        result["by_status"][o.status]["count"] += 1
        result["by_status"][o.status]["amount"] += o.total_amount
    return result


@router.get("/export/excel")
def export_excel(
    period: str = Query(default="month"),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    status: str | None = Query(default=None),
    search: str | None = Query(default=None),
    db: Session = Depends(get_db),
):
    """خروجی اکسل از سفارش‌ها"""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    query = db.query(models.Order)
    start, end = _date_range(period, date_from, date_to)
    if start:
        query = query.filter(models.Order.created_at >= start)
    if end:
        query = query.filter(models.Order.created_at <= end)
    if status:
        query = query.filter(models.Order.status == status)
    if search:
        s = f"%{search}%"
        query = query.filter(
            models.Order.code.ilike(s) | models.Order.customer_name.ilike(s)
        )
    orders = query.order_by(models.Order.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "سفارش‌ها"
    ws.sheet_view.rightToLeft = True

    headers = ["ردیف", "کد سفارش", "نام مشتری", "وضعیت", "منبع", "مبلغ (تومان)", "تاریخ ثبت", "اقلام"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    status_map = {
        "pending": "در انتظار", "preparing": "در حال آماده‌سازی",
        "ready": "آماده", "completed": "تکمیل شده", "cancelled": "لغو شده"
    }
    source_map = {"online": "آنلاین", "walk_in": "حضوری"}

    for row_idx, o in enumerate(orders, 2):
        items_str = "، ".join(f"{it.product_name}×{it.quantity}" for it in o.items)
        ws.append([
            row_idx - 1,
            o.code,
            o.customer_name or "",
            status_map.get(o.status, o.status),
            source_map.get(o.source, o.source),
            o.total_amount,
            o.created_at.strftime("%Y-%m-%d %H:%M"),
            items_str,
        ])

    col_widths = [8, 18, 20, 18, 12, 18, 20, 50]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"orders_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export/pdf")
def export_pdf(
    period: str = Query(default="month"),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    status: str | None = Query(default=None),
    search: str | None = Query(default=None),
    db: Session = Depends(get_db),
):
    """خروجی PDF از سفارش‌ها"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    query = db.query(models.Order)
    start, end = _date_range(period, date_from, date_to)
    if start:
        query = query.filter(models.Order.created_at >= start)
    if end:
        query = query.filter(models.Order.created_at <= end)
    if status:
        query = query.filter(models.Order.status == status)
    if search:
        s = f"%{search}%"
        query = query.filter(
            models.Order.code.ilike(s) | models.Order.customer_name.ilike(s)
        )
    orders = query.order_by(models.Order.created_at.desc()).all()

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm, topMargin=1.5*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()

    status_map = {
        "pending": "Pending", "preparing": "Preparing",
        "ready": "Ready", "completed": "Completed", "cancelled": "Cancelled"
    }

    data = [["#", "Code", "Customer", "Status", "Source", "Amount", "Date", "Items"]]
    for i, o in enumerate(orders, 1):
        items_str = ", ".join(f"{it.product_name}x{it.quantity}" for it in o.items)
        data.append([
            str(i), o.code, o.customer_name or "-",
            status_map.get(o.status, o.status),
            o.source, f"{o.total_amount:,}",
            o.created_at.strftime("%Y-%m-%d %H:%M"),
            items_str[:60] + ("..." if len(items_str) > 60 else ""),
        ])

    t = Table(data, colWidths=[1*cm, 3*cm, 4*cm, 3*cm, 2.5*cm, 3*cm, 4*cm, None])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF2FF")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))

    title = Paragraph(f"<b>Orders Report — {datetime.now().strftime('%Y-%m-%d')}</b>", styles["Title"])
    elements = [title, Spacer(1, 0.5*cm), t]
    doc.build(elements)

    buf.seek(0)
    filename = f"orders_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------- داشبورد ----------------


@router.get("/dashboard/today")
def dashboard_today(db: Session = Depends(get_db)):
    """آمار امروز: تعداد سفارش، درآمد، میانگین مبلغ"""
    start_of_today = datetime.combine(datetime.now().date(), time.min)
    end_of_today = datetime.combine(datetime.now().date(), time.max)
    orders = (
        db.query(models.Order)
        .filter(models.Order.created_at >= start_of_today)
        .filter(models.Order.created_at <= end_of_today)
        .all()
    )
    completed = [o for o in orders if o.status == "completed"]
    total_revenue = sum(o.total_amount for o in completed)
    avg_amount = total_revenue // len(completed) if completed else 0
    return {
        "total_orders": len(orders),
        "completed_orders": len(completed),
        "total_revenue": total_revenue,
        "avg_amount": avg_amount,
    }


@router.get("/dashboard/revenue-week")
def dashboard_revenue_week(db: Session = Depends(get_db)):
    """درآمد ۷ روز اخیر (روزانه)"""
    result = []
    for i in range(6, -1, -1):
        day = datetime.now().date() - timedelta(days=i)
        start = datetime.combine(day, time.min)
        end = datetime.combine(day, time.max)
        orders = (
            db.query(models.Order)
            .filter(models.Order.created_at >= start)
            .filter(models.Order.created_at <= end)
            .filter(models.Order.status == "completed")
            .all()
        )
        revenue = sum(o.total_amount for o in orders)
        result.append({
            "date": day.strftime("%Y-%m-%d"),
            "label": day.strftime("%m/%d"),
            "revenue": revenue,
            "count": len(orders),
        })
    return result


@router.get("/dashboard/top-products")
def dashboard_top_products(db: Session = Depends(get_db)):
    """۵ محصول پرفروش (بر اساس تعداد فروخته‌شده)"""
    from sqlalchemy import func

    start_of_today = datetime.combine(datetime.now().date(), time.min)
    results = (
        db.query(
            models.OrderItem.product_name,
            func.sum(models.OrderItem.quantity).label("total_qty"),
            func.sum(models.OrderItem.unit_price * models.OrderItem.quantity).label("total_revenue"),
        )
        .join(models.Order, models.OrderItem.order_id == models.Order.id)
        .filter(models.Order.created_at >= start_of_today)
        .filter(models.Order.status.in_(["completed", "ready", "preparing"]))
        .group_by(models.OrderItem.product_name)
        .order_by(func.sum(models.OrderItem.quantity).desc())
        .limit(5)
        .all()
    )
    return [
        {"name": r.product_name, "quantity": r.total_qty, "revenue": r.total_revenue}
        for r in results
    ]


@router.get("/dashboard/recent-orders")
def dashboard_recent_orders(db: Session = Depends(get_db)):
    """۱۰ سفارش اخیر"""
    orders = (
        db.query(models.Order)
        .order_by(models.Order.created_at.desc())
        .limit(10)
        .all()
    )
    return [order_to_out(o, with_qr=False) for o in orders]
